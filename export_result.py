
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

# 重新运行分析获取数据
xlsx = pd.ExcelFile('/root/.openclaw/media/inbound/ç´_æ1---cce5ba51-3360-4b9c-88ee-1fff7862b38f.xlsx')
df = pd.read_excel(xlsx, sheet_name='Sheet1', header=None)

# 提取数据
years_data = []
current_year = None

for idx, row in df.iterrows():
    name = str(row[0]) if pd.notna(row[0]) else ''
    if '四分位数' in name or '中位数' in name:
        if '上四分位数' in name:
            if current_year is not None:
                years_data.append(current_year)
            year_name = name.split('年')[0] + '年'
            current_year = {
                'year': year_name,
                'A': {'q3': None, 'q2': None, 'q1': None},
                'B': {'q3': None, 'q2': None, 'q1': None},
                'C': {'q3': None, 'q2': None, 'q1': None},
                'enterprises': []
            }
            for col in [1, 2, 3]:
                if pd.notna(row[col]):
                    current_year[['A', 'B', 'C'][col-1]]['q3'] = float(row[col])
        elif '中位数' in name:
            for col in [1, 2, 3]:
                if pd.notna(row[col]):
                    current_year[['A', 'B', 'C'][col-1]]['q2'] = float(row[col])
        elif '下四分位数' in name:
            for col in [1, 2, 3]:
                if pd.notna(row[col]):
                    current_year[['A', 'B', 'C'][col-1]]['q1'] = float(row[col])
    elif name.startswith('企业') and current_year is not None:
        ent = {}
        for col in [1, 2, 3]:
            if pd.notna(row[col]) and str(row[col]).strip() != '':
                ent[['A', 'B', 'C'][col-1]] = float(row[col])
            else:
                ent[['A', 'B', 'C'][col-1]] = None
        current_year['enterprises'].append(ent)

if current_year is not None:
    years_data.append(current_year)

def count_intervals(year_data, indicator):
    q3 = year_data[indicator]['q3']
    q2 = year_data[indicator]['q2']
    q1 = year_data[indicator]['q1']
    
    result = {
        '≥上四分位': 0,
        '≥中位数且＜上四分位': 0,
        '≥下四分位且＜中位数': 0,
        '＜下四分位': 0
    }
    
    for ent in year_data['enterprises']:
        val = ent[indicator]
        if val is None:
            continue
        if val >= q3:
            result['≥上四分位'] += 1
        elif val >= q2:
            result['≥中位数且＜上四分位'] += 1
        elif val >= q1:
            result['≥下四分位且＜中位数'] += 1
        else:
            result['＜下四分位'] += 1
    
    return result

# 创建工作簿
wb = Workbook()
ws1 = wb.active
ws1.title = '汇总结果'

# 设置样式
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
header_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center')

# 写入汇总表标题
ws1['A1'] = '区间'
ws1['B1'] = '指标 A'
ws1['C1'] = '指标 B'
ws1['D1'] = '指标 C'

for cell in ['A1', 'B1', 'C1', 'D1']:
    ws1[cell].font = header_font
    ws1[cell].alignment = center_align
    ws1[cell].border = thin_border

# 汇总数据
total_all = {
    'A': {'≥上四分位': 0, '≥中位数且＜上四分位': 0, '≥下四分位且＜中位数': 0, '＜下四分位': 0},
    'B': {'≥上四分位': 0, '≥中位数且＜上四分位': 0, '≥下四分位且＜中位数': 0, '＜下四分位': 0},
    'C': {'≥上四分位': 0, '≥中位数且＜上四分位': 0, '≥下四分位且＜中位数': 0, '＜下四分位': 0}
}

for yd in years_data:
    for ind in ['A', 'B', 'C']:
        res = count_intervals(yd, ind)
        for k in res:
            total_all[ind][k] += res[k]

# 写入汇总数据
intervals = ['≥上四分位', '≥中位数且＜上四分位', '≥下四分位且＜中位数', '＜下四分位']
row_num = 2
for interval in intervals:
    ws1.cell(row=row_num, column=1, value=interval).border = thin_border
    ws1.cell(row=row_num, column=1).alignment = center_align
    ws1.cell(row=row_num, column=2, value=total_all['A'][interval]).border = thin_border
    ws1.cell(row=row_num, column=2).alignment = center_align
    ws1.cell(row=row_num, column=3, value=total_all['B'][interval]).border = thin_border
    ws1.cell(row=row_num, column=3).alignment = center_align
    ws1.cell(row=row_num, column=4, value=total_all['C'][interval]).border = thin_border
    ws1.cell(row=row_num, column=4).alignment = center_align
    row_num += 1

# 写入合计行
ws1.cell(row=row_num, column=1, value='合计').font = header_font
ws1.cell(row=row_num, column=1).border = thin_border
ws1.cell(row=row_num, column=1).alignment = center_align
ws1.cell(row=row_num, column=2, value=sum(total_all['A'].values())).font = header_font
ws1.cell(row=row_num, column=2).border = thin_border
ws1.cell(row=row_num, column=2).alignment = center_align
ws1.cell(row=row_num, column=3, value=sum(total_all['B'].values())).font = header_font
ws1.cell(row=row_num, column=3).border = thin_border
ws1.cell(row=row_num, column=3).alignment = center_align
ws1.cell(row=row_num, column=4, value=sum(total_all['C'].values())).font = header_font
ws1.cell(row=row_num, column=4).border = thin_border
ws1.cell(row=row_num, column=4).alignment = center_align

# 创建第二个工作表 - 各年份详细统计
ws2 = wb.create_sheet(title='分年份统计')

headers = ['年份', '企业数', '指标', '≥上四分位', '≥中位数且＜上四分位', '≥下四分位且＜中位数', '＜下四分位', '合计']
for col_num, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col_num, value=header)
    cell.font = header_font
    cell.alignment = center_align
    cell.border = thin_border

row_num = 2
for yd in years_data:
    for ind in ['A', 'B', 'C']:
        res = count_intervals(yd, ind)
        total = sum(res.values())
        ws2.cell(row=row_num, column=1, value=yd['year']).border = thin_border
        ws2.cell(row=row_num, column=1).alignment = center_align
        ws2.cell(row=row_num, column=2, value=len(yd['enterprises'])).border = thin_border
        ws2.cell(row=row_num, column=2).alignment = center_align
        ws2.cell(row=row_num, column=3, value='指标 ' + ind).border = thin_border
        ws2.cell(row=row_num, column=3).alignment = center_align
        col_num = 4
        for interval in intervals:
            ws2.cell(row=row_num, column=col_num, value=res[interval]).border = thin_border
            ws2.cell(row=row_num, column=col_num).alignment = center_align
            col_num += 1
        ws2.cell(row=row_num, column=col_num, value=total).border = thin_border
        ws2.cell(row=row_num, column=col_num).alignment = center_align
        row_num += 1

# 设置列宽
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 10
ws1.column_dimensions['C'].width = 10
ws1.column_dimensions['D'].width = 10

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 8
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 22
ws2.column_dimensions['F'].width = 22
ws2.column_dimensions['G'].width = 14
ws2.column_dimensions['H'].width = 8

# 保存文件
output_path = '/home/liam/.openclaw/workspace/统计结果.xlsx'
wb.save(output_path)
print(f"文件已保存到: {output_path}")
